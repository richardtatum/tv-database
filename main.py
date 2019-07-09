from bs4 import BeautifulSoup as bs
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.cell.cell import Cell
import logging
from services import EmailConnect, Dropbox
import requests
import os
import time
from logging.handlers import TimedRotatingFileHandler

LOCAL_FILE = 'data/International Format Tracker.xlsx'
REMOTE_FILE = '/International Format Tracker.xlsx'

# Logging Handler
log_format = logging.Formatter('%(asctime)s | %(name)s %(levelname)s: %(message)s', '%Y-%m-%d %H:%M:%S')
handler = TimedRotatingFileHandler(
    'logs/database.log',
    when='D',
    interval=7,
    backupCount=10,
)
handler.setFormatter(log_format)
logger = logging.getLogger()
logger.addHandler(handler)
logger.setLevel(logging.INFO)


# Applies formatting to the data to match the sheet
def apply_formatting(data):
    logger.info('Matching formatting to worksheet')
    for d in data:
        d = Cell(ws, column='A', row=1, value=d)
        d.font = Font(name='Calibri', size=9)
        d.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        yield d


# Sets column C to justify left
def final_formatting():
    logger.info('Applying final formatting')
    column_c = ws['C']
    for cell in column_c:
        cell.alignment = Alignment(
            horizontal='left',
            vertical='center',
            wrap_text=True
        )


def parse_content(url):
    r = requests.get(url)
    soup = bs(r.text, 'html.parser')
    data = []

    # Section that holds the date and country
    link_data = soup.find('div', class_='profile_tweet_date_country').get_text().split()

    # Format the date to requested style
    try:
        time = datetime.strptime(link_data[0], '%d/%m/%Y').strftime('%d/%m/%y')
    except ValueError:
        time = date.today().strftime('%d/%m/%y')

    # Appending data to a list (required for openpyxl.append())
    data.append(time)
    data.append(' '.join(link_data[2:]))
    data.append(soup.find('div', class_='profile_tweet_content').get_text())

    # Pass data to worksheet
    ws.append(apply_formatting(data))


# Removes duplicates whilst maintaining order
def remove_duplicates(list_w_duplicates):
    app.logger('Removing duplicates')
    return list(dict.fromkeys(list_w_duplicates))


# Picks out the correct content links from the email
def acquire_links(subject):
    # Acquire the ID of the email
    ids = gmail.get_id_by_subject(subject)

    # If no ID's are returned, logout and exit
    if len(ids) < 1:
        logger.info(f'No emails found matching the requested subject <{subject}>.')
        logger.info('Exiting...')
        gmail.logout()
        exit()

    # If there are links, pull the raw HTML
    html = gmail.get_html(ids)

    links = []
    for email in html:
        e_soup = bs(email, 'html.parser')

        # Adds correct links to a masterlist
        for link in e_soup.find_all('a'):
            if 'tvbizz.net/newsitemsocial' in link.get('href'):
                links.append(link.get('href'))

    # Move the email once we have finished
    gmail.move(ids, '[Google Mail]/All Mail')

    return remove_duplicates(links)


if __name__ == '__main__':
    # Create a gmail instance and login
    gmail = EmailConnect(
        os.getenv('IMAP_HOST'),
        os.getenv('IMAP_USER'),
        os.getenv('APP_PASS'),
    )

    # Same for Dropbox
    box = Dropbox(os.getenv('DBX_TOKEN'))
    box.download(LOCAL_FILE, REMOTE_FILE)

    # Load the document and worksheet
    wb = load_workbook(LOCAL_FILE)
    ws = wb['TV BIZZ']
    logger.info('Loading file.')

    # Aquire all the data from each link and add it to the file
    link_list = acquire_links('Latest headlines on TVBIZZ')
    for url in link_list:
        parse_content(url)
    logger.info(f'Added data from {len(link_list)} links.')

    # Terminate gmail connection
    gmail.logout()

    # Apply formatting
    final_formatting()

    # Save the file
    wb.save(LOCAL_FILE)
    logger.info('File saved.')

    # Sleep to make sure file is correctly saved before uploading
    time.sleep(0.5)

    # Upload to dropbox
    box.upload(LOCAL_FILE, REMOTE_FILE)
